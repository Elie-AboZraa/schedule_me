import re
import sys
import csv
from openpyxl import load_workbook,Workbook

# Helper functions to identify patterns
def is_day(cell_value):
    return re.match(r"^(Saturday|Sunday|Monday|Tuesday)$", str(cell_value))

def is_location(cell_value):
    return re.match(r"^\d+.", str(cell_value))

def is_subject( ws, col_index, row_index):
    cell = ws.cell(row=row_index + 1, column=col_index + 1)
    if cell.value is None :
        return False
    if find_teacher(ws, col_index, row_index):
        return True
    if  cell.fill.start_color.index == 'FF2A6099' or cell.fill.start_color.index=='FF158466':
        return True
    # if the before failed fail back to the legacy way (un acurate way)
    #cell_value = str(cell_value).replace(' ', '')
    return re.match(r'[\w+\s]+\d*-*ش-*\d', str(cell.value))

def is_teacher(cell_value):
    return re.match(r"^د\.\s*\w+|^أ\.\s*\w+|^م\.\s*\w+", str(cell_value))

def is_time_range(cell_value):
    return re.match(r"^\d+:\d+-\d+:\d+$", str(cell_value))

def find_teacher(ws, col_index, row_index):
    if row_index + 1 < ws.max_row:
        cell = ws.cell(row=row_index + 2, column=col_index + 1).value
        if is_teacher(cell):
            return cell

def find_location(ws, col_index, row_index):
    for i in reversed(range(col_index + 1)):
        if col_index - i + 1 <= 0:
            continue
        cell = ws.cell(row=row_index + 1, column=col_index - i + 1).value
        if is_location(cell):
            return cell

def find_time(ws, col_index, row_index):
    for i in reversed(range(row_index + 1)):
        cell = ws.cell(row=row_index - i + 1, column=col_index + 1).value
        if is_time_range(cell):
            #08:00-12:00
            return cell

def find_day(ws, col_index, row_index):
    day_count = 0
    for i in reversed(range(col_index + 1)):
        cell = ws.cell(row=row_index + 1, column=col_index - i + 1).value
        if is_location(cell):
            day_count += 1
    days = {1: "Saturday", 2: "Sunday", 3: "Monday", 4: "Tuesday", 5: "Wednesday", 6: "Thursday"}
    return days.get(day_count, "Unknown")
def  find_lecturetype(ws, col_index, row_index):
    cell = ws.cell(row=row_index + 1, column=col_index + 1)
    if cell.fill.start_color.index == 'FF2A6099':
        return "T"
    else :
        return "P"
    
def after_clean(subject_name:str):
    subject_name=subject_name.strip()
    match=re.match(r".+\s\d",subject_name)
    if match:
        match= match.group(0)
        # removing the space before the diget to uniform the subjects name , it can happen two subject that are the same but defrent name so it will be treated as a difrent subject , we will try to ellimnate that
        return match[:-2]+match[-1]
    return subject_name

def calculate_Subject_Hours(Lectures:list):
    Theory_Time=0
    Lap_Time=0
    for l in Lectures:
        if l.get("RepitionTime")=='1':
            if l.get("LectureType")=='T':
                Theory_Time+=1
            else :
                Lap_Time+=1
    if (Lap_Time+Theory_Time)==0:
        return "None"
    return Lap_Time+Theory_Time+1

def _clean_subject_name(subject):

    try:
        tmp=None
        subject = str(subject)
        if re.search(r".+\d\/.+\d", subject):
            tmp=subject.split("/")
            return [_clean_subject_name(tmp[0])[0],_clean_subject_name(tmp[1])[0]]
        if re.search(r".+\d\+ش\d+", subject):
            tmp=subject.split("+")
            tmp1=tmp[0][:-1]+tmp[1][-1]
            return [_clean_subject_name(tmp[0])[0],_clean_subject_name(tmp1)[0]]
        if re.search(r".+ش\d+", subject):
            match = re.search(r".+ش\d+", subject)
            subject = match.group(0)
        if re.search(r".+ش\s\d+", subject):
            match = re.search(r".+ش\s\d+", subject)
            subject = match.group(0)
        elif re.search(r".+ش-\d+", subject):
            match = re.search(r".+ش-\d+", subject)
            subject = match.group(0)

        if "- ش" in subject:
            tmp = subject.split("- ش")
            tmp[0]=after_clean(tmp[0])
            return [[tmp[0].strip(), tmp[1].strip()]]
        elif " -ش" in subject:
            tmp = subject.split(" -ش")
            tmp[0]=after_clean(tmp[0])
            return [[tmp[0].strip(), tmp[1].strip()]]
        elif "-ش" in subject:
            tmp = subject.split("-ش")
            tmp[0]=after_clean(tmp[0])
            return [[tmp[0].strip(), tmp[1].strip()]]
        elif "ش-" in subject:
            tmp = subject.split("ش-")
            tmp[0]=after_clean(tmp[0])
            return [[tmp[0].strip(), tmp[1].strip()]]
        elif re.match(r".+ش\s\d", subject):
            tmp = subject.split(" ش ")
            tmp[0]=after_clean(tmp[0])
            return [[tmp[0].strip(), tmp[1].strip()]]
        elif re.match(r".+ش\d", subject):
            key_part = subject[:-2].strip()
            value_part = subject[-1].strip()
            
            return [[after_clean(key_part), value_part]]
        return [[subject.strip(), "1"]]
    except Exception as e:
        raise ValueError(f"Error cleaning subject names: {e}")


def main():
    # Path to your Excel file
    file_path = sys.path[0] + "/2nd_semester12-03-2025.xlsx"

    # Load the workbook and select the first sheet
    wb = load_workbook(filename=file_path)
    ws = wb.active

    # Initialize storage for extracted data
    Subjects = {}
    Lecture={}

    # Loop through each column
    lectureid=0
    for col_index in range(ws.max_column):
        for row_index in range(ws.max_row):
            cell = ws.cell(row=row_index + 1, column=col_index + 1).value

            if None is cell:
                pass

            cell_value = str(cell)
            Lecture = {}
            # if the cell is one of those(day-location-teacher-time) then ignore it and go to the next cell 
            # thats becouse its nearly imposable to identify subject from normal text 
            if is_day(cell_value) or is_location(cell_value) or is_teacher(cell_value) or is_time_range(cell_value):
                pass
            elif is_subject(ws, col_index, row_index):
                tmp=None
                for tmp in _clean_subject_name(cell_value):
                    add=True
                    subject_key = tmp[0]
                    rep_time = tmp[1]
                    if subject_key not in Subjects:
                        Subjects[subject_key] = []
                    Lecture["Teacher"] = find_teacher(ws, col_index, row_index)
                    '''time_tmp=None
                    time_tmp=find_time(ws, col_index, row_index)
                    Lecture["Timerange-start"] =time_tmp[0]
                    Lecture["Timerange-end"] = time_tmp[1]'''
                    Lecture["Timerange"]=[find_time(ws, col_index, row_index)]
                    tsad=find_location(ws, col_index, row_index).__str__()
                    Lecture["Location"] = tsad.__str__()
                    Lecture["RepitionTime"] = rep_time
                    Lecture["Day"] = find_day(ws, col_index, row_index)
                    # the LectureType reffure to if the lecture is therotical or in labs 
                    Lecture["LectureType"] = find_lecturetype(ws, col_index, row_index)
                    Lecture["LectureId"] = lectureid.__str__()
                    
                    for lecture in Subjects[subject_key]:
                        if lecture["LectureType"]==Lecture["LectureType"] and lecture["RepitionTime"]== rep_time:
                            lecture["Timerange"].append(Lecture["Timerange"][0])
                            add=False
                    if add:
                        lectureid+=1
                        Subjects[subject_key].append(Lecture.copy())



    # Create a new workbook and select the active worksheet
    #wb_out = Workbook()
    #ws_out = wb_out.active
    #print(Subjects)
    #{sub_name:[{Lectures},{Lectures},{Lectures}]}
    # Write data rows
    open(sys.path[0] + "/new-extracted_schedule_openpyxl.csv", 'w', newline='').close()
    with open(sys.path[0] + "/new-extracted_schedule_openpyxl.csv", 'a', newline='') as csvfile:
        
        csv_writer = csv.writer(csvfile)
        for subject_name, Lectures in Subjects.items():
            for dicks in Lectures:
                csv_writer.writerow(
                [
                    dicks.get("LectureId", ""),
                    subject_name,
                    dicks.get("RepitionTime", ""),
                    dicks.get("Day", ""),
                    dicks.get("Timerange", ""),
                    dicks.get("Location", ""),
                    dicks.get("Teacher", ""),
                    dicks.get("LectureType", ""),

                ]


                )
                #ws_out.append(row)
            csv_writer.writerow(
            ["Sub",subject_name,calculate_Subject_Hours(Lectures)]
            )

    # Save the workbook
    #output_path = sys.path[0] + "/extracted_schedule_openpyxl.xlsx"
    #wb_out.save(output_path)

if __name__ == "__main__":
    main()